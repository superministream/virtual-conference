import openpyxl

def open(filename):
    db = ExcelDb(db=openpyxl.load_workbook(filename=filename, data_only=True))
    return db

# It's nice to have an Excel file since we can paste it out to Google sheets as
# a convenient "export". This implements a sort of database interface on top
# of the Excel sheet to work with it on the Python side
class ExcelDb:
    def __init__(self, db=None):
        if db == None:
            self.db = openpyxl.Workbook()
        else:
            self.db = db

    def save(self, filename):
        self.db.save(filename)

    # Get a sheet (or "table")
    def get_table(self, table_name):
        table = ExcelTable(self.db, self.db[table_name])
        table.load_index()
        return table

    def create_table(self, table, index):
        table = ExcelTable(self.db, self.db.create_sheet(title=table))
        table.set_index(index)
        return table

    def table_names(self):
        return self.db.get_sheet_names()

class ExcelTable:
    def __init__(self, db, table):
        self.db = db
        self.table = table

    # Index should be an array of column names in the order desired 
    def set_index(self, index):
        header_style = "40 % - Accent1"
        self.index = {}
        for col, name in enumerate(index):
            self.index[name] = col + 1
            self.entry(1, col + 1).value = name
            self.entry(1, col + 1).style = header_style

    def load_index(self):
        self.index = {}
        for c in self.table.iter_cols(max_row=1):
            # Skip empty columns
            if not c[0].value:
                continue
            self.index[c[0].value] = c[0].column

    def entry(self, row, col):
        if type(col) is int:
            return self.table.cell(row, col)
        return self.entry(row, self.index[col])

    # Find all rows with attribute matching value
    def find(self, attrib, value):
        found = []
        col = self.index[attrib]
        for r in self.table.iter_rows(min_row=2, min_col=col, max_col=col):
            if r[0].value == value:
                found.append(r[0].row)
        return found

    # Find all rows where the function returns true. The function
    # is passed the row being evaluated
    def find_if(self, fcn):
        found = []
        for r in range(2, self.table.max_row + 1):
            if fcn(self.row(r)):
                found.append(r)
        return found

    def row(self, row):
        data = {}
        for k, v in self.index.items():
            data[k] = self.entry(row, v)
        return data

    def items(self):
        items = []
        for r in range(2, self.table.max_row + 1):
            items.append(self.row(r))
        return items

    def write_row(self, row, data):
        for k, v in data.items():
            self.entry(row, self.index[k]).value = v

    # Append a dict of data into the database, mapping the keys
    # to the corresponding output column
    def append_row(self, data):
        row_array = [None] * len(self.index)
        for k, v in data.items():
            row_array[self.index[k] - 1] = v
        self.table.append(row_array)
        return self.table.max_row
